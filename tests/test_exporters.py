"""
Tests for exporters.py — CSVExporter, JSONExporter, ExcelExporter.

All tests use synthetic BatchResult objects; no image processing required.
"""

import csv
import json
import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from data_models import (
    BoundingBox,
    ElementType,
    IndexEntry,
    TableCell,
    TableStructure,
    TableOfContents,
    Watermark,
    FigureRegion,
    Reference,
    CodeBlock,
    ListItem,
    ListStructure,
)
from exporters import CSVExporter, ExporterConfig, JSONExporter
from helpers import make_batch, make_doc, make_element, make_bbox


# ── Shared fixtures ────────────────────────────────────────────────────────────

def _table_elem():
    bbox = make_bbox(0, 0, 200, 100)
    cells = [
        TableCell("H1", 0, 0, BoundingBox(0, 0, 50, 30), 0.9, is_header=True),
        TableCell("H2", 0, 1, BoundingBox(50, 0, 100, 30), 0.9, is_header=True),
        TableCell("D1", 1, 0, BoundingBox(0, 30, 50, 60), 0.8),
        TableCell("D2", 1, 1, BoundingBox(50, 30, 100, 60), 0.8),
    ]
    table = TableStructure(cells, BoundingBox(0, 0, 200, 100), 0.85)
    return make_element(ElementType.TABLE, table, element_id="tbl1")


def _list_elem():
    items = [
        ListItem("First item", 0, make_bbox(), 0.9),
        ListItem("Second item", 0, make_bbox(), 0.9),
        ListItem("  Sub item", 1, make_bbox(), 0.85),
    ]
    lst = ListStructure(items, [], make_bbox(0, 0, 300, 80), 0.87)
    return make_element(ElementType.LIST, lst, element_id="lst1")


def _watermark_elem():
    wm = Watermark("DRAFT", make_bbox(0, 0, 400, 50), 0.75, opacity_estimate=0.3)
    return make_element(ElementType.WATERMARK, wm, element_id="wm1")


def _figure_elem():
    fig = FigureRegion(make_bbox(0, 0, 300, 200), 0.80, figure_type="chart")
    return make_element(ElementType.FIGURE, fig, element_id="fig1")


def _ref_elem():
    ref = Reference("[1] Author (2020). Title.", "bibliography", "1", make_bbox())
    return make_element(ElementType.REFERENCE, ref, element_id="ref1")


def _code_elem():
    cb = CodeBlock("def foo(): pass", make_bbox(), 0.88, language="python")
    return make_element(ElementType.CODE_BLOCK, cb, element_id="cb1")


def _toc_elem():
    toc = TableOfContents("Introduction", 1, 1, make_bbox(), 0.9)
    return make_element(ElementType.TABLE_OF_CONTENTS, toc, element_id="toc1")


def _index_elem():
    ie = IndexEntry("algorithm", [5, 10], 1, make_bbox(), 0.8)
    return make_element(ElementType.INDEX, ie, element_id="idx1")


# ── CSVExporter ────────────────────────────────────────────────────────────────

class TestCSVExporter:
    def test_empty_batch_writes_empty_file(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[]])
        CSVExporter().export(batch, path)
        assert path.exists()
        assert path.stat().st_size == 0

    def test_row_count_matches_elements(self, tmp_path):
        path = tmp_path / "out.csv"
        elems = [
            make_element(ElementType.TEXT, "para 1"),
            make_element(ElementType.TEXT, "para 2"),
            make_element(ElementType.HEADING, "Title"),
        ]
        batch = make_batch([elems])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 3

    def test_common_columns_present(self, tmp_path):
        path = tmp_path / "out.csv"
        elems = [make_element(ElementType.TEXT, "hello", element_id="e1")]
        batch = make_batch([elems])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            header = reader.fieldnames
        required = {
            "batch_id", "document_id", "element_id", "element_type",
            "page", "confidence", "content_text",
            "bbox_x_min", "bbox_y_min", "bbox_x_max", "bbox_y_max",
        }
        assert required.issubset(set(header))

    def test_type_columns_present_when_enabled(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[make_element(ElementType.TEXT, "x")]])
        CSVExporter(ExporterConfig(include_type_columns=True)).export(batch, path)
        with open(path, encoding="utf-8") as f:
            header = csv.DictReader(f).fieldnames
        assert "table_rows" in header
        assert "formula_latex" in header

    def test_type_columns_absent_when_disabled(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[make_element(ElementType.TEXT, "x")]])
        CSVExporter(ExporterConfig(include_type_columns=False)).export(batch, path)
        with open(path, encoding="utf-8") as f:
            header = csv.DictReader(f).fieldnames
        assert "table_rows" not in header

    def test_table_type_columns_populated(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[_table_elem()]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert row["table_rows"] == "2"
        assert row["table_cols"] == "2"

    def test_watermark_type_columns_populated(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[_watermark_elem()]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert row["watermark_opacity"] != ""

    def test_reference_type_columns_populated(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[_ref_elem()]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert row["ref_type"] == "bibliography"
        assert row["ref_id"] == "1"

    def test_code_language_column(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[_code_elem()]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert row["code_language"] == "python"

    def test_toc_columns_populated(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[_toc_elem()]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert row["toc_page_ref"] == "1"
        assert row["toc_level"] == "1"

    def test_index_columns_populated(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[_index_elem()]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert row["index_term"] == "algorithm"
        assert "5" in row["index_pages"]

    def test_multi_document_batch(self, tmp_path):
        path = tmp_path / "out.csv"
        elems_a = [make_element(ElementType.TEXT, "doc A text")]
        elems_b = [make_element(ElementType.HEADING, "Doc B heading")]
        batch = make_batch([elems_a, elems_b])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2
        doc_ids = {r["document_id"] for r in rows}
        assert len(doc_ids) == 2

    def test_export_document_convenience(self, tmp_path):
        path = tmp_path / "doc.csv"
        doc = make_doc([make_element(ElementType.TEXT, "hello")])
        CSVExporter().export_document(doc, path)
        with open(path, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 1

    def test_content_text_for_str_content(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[make_element(ElementType.TEXT, "hello world")]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert row["content_text"] == "hello world"

    def test_content_text_for_table(self, tmp_path):
        path = tmp_path / "out.csv"
        batch = make_batch([[_table_elem()]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            row = list(csv.DictReader(f))[0]
        assert "Table" in row["content_text"]

    def test_parent_id_and_child_count(self, tmp_path):
        path = tmp_path / "out.csv"
        parent = make_element(ElementType.HEADING, "H1", element_id="p")
        child = make_element(ElementType.TEXT, "body", element_id="c", parent_id="p")
        parent.child_ids = ["c"]
        batch = make_batch([[parent, child]])
        CSVExporter().export(batch, path)
        with open(path, encoding="utf-8") as f:
            rows = {r["element_id"]: r for r in csv.DictReader(f)}
        assert rows["c"]["parent_id"] == "p"
        assert rows["p"]["child_count"] == "1"

    def test_creates_parent_directory(self, tmp_path):
        deep_path = tmp_path / "a" / "b" / "out.csv"
        batch = make_batch([[make_element(ElementType.TEXT, "x")]])
        CSVExporter().export(batch, deep_path)
        assert deep_path.exists()


# ── JSONExporter ───────────────────────────────────────────────────────────────

class TestJSONExporter:
    def test_valid_json_output(self, tmp_path):
        path = tmp_path / "out.json"
        batch = make_batch([[make_element(ElementType.TEXT, "hello")]])
        JSONExporter().export(batch, path)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert "batch_id" in data
        assert "documents" in data

    def test_document_fields(self, tmp_path):
        path = tmp_path / "out.json"
        batch = make_batch([[make_element(ElementType.TEXT, "hello")]])
        JSONExporter().export(batch, path)
        doc = json.loads(path.read_text(encoding="utf-8"))["documents"][0]
        assert "document_id" in doc
        assert "elements" in doc
        assert "processing_status" in doc

    def test_hierarchical_mode_embeds_children(self, tmp_path):
        path = tmp_path / "out.json"
        parent = make_element(ElementType.HEADING, "H1", element_id="p")
        child = make_element(ElementType.TEXT, "body", element_id="c", parent_id="p")
        parent.child_ids = ["c"]
        batch = make_batch([[parent, child]])
        JSONExporter(ExporterConfig(json_hierarchical=True)).export(batch, path)
        doc = json.loads(path.read_text(encoding="utf-8"))["documents"][0]
        roots = doc["elements"]
        # In hierarchical mode the child should be nested inside the parent, not at root
        root_ids = [e["element_id"] for e in roots]
        assert "p" in root_ids
        # Child should appear inside parent's children list
        parent_elem = next(e for e in roots if e["element_id"] == "p")
        assert any(ch["element_id"] == "c" for ch in parent_elem.get("children", []))

    def test_flat_mode_all_elements_at_root(self, tmp_path):
        path = tmp_path / "out.json"
        parent = make_element(ElementType.HEADING, "H1", element_id="p")
        child = make_element(ElementType.TEXT, "body", element_id="c", parent_id="p")
        parent.child_ids = ["c"]
        batch = make_batch([[parent, child]])
        JSONExporter(ExporterConfig(json_hierarchical=False)).export(batch, path)
        doc = json.loads(path.read_text(encoding="utf-8"))["documents"][0]
        elem_ids = {e["element_id"] for e in doc["elements"]}
        assert "p" in elem_ids and "c" in elem_ids

    def test_statistics_present(self, tmp_path):
        path = tmp_path / "out.json"
        batch = make_batch([[make_element(ElementType.TEXT, "x")]])
        JSONExporter().export(batch, path)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert "statistics" in data
        assert "total_elements" in data["statistics"]

    def test_empty_batch(self, tmp_path):
        path = tmp_path / "out.json"
        batch = make_batch([[]])
        JSONExporter().export(batch, path)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert data["documents"][0]["elements"] == []

    def test_compact_mode(self, tmp_path):
        path = tmp_path / "out.json"
        batch = make_batch([[make_element(ElementType.TEXT, "x")]])
        JSONExporter(ExporterConfig(json_indent=None)).export(batch, path)
        raw = path.read_text(encoding="utf-8")
        assert "\n" not in raw  # compact JSON has no newlines

    def test_type_detail_present_when_enabled(self, tmp_path):
        path = tmp_path / "out.json"
        batch = make_batch([[_table_elem()]])
        JSONExporter(ExporterConfig(include_type_columns=True)).export(batch, path)
        doc = json.loads(path.read_text(encoding="utf-8"))["documents"][0]
        elem = doc["elements"][0]
        # In hierarchical mode, table has no parent so it's at root
        if "children" in elem:  # hierarchical root element
            assert "type_detail" in elem

    def test_export_document_convenience(self, tmp_path):
        path = tmp_path / "doc.json"
        doc = make_doc([make_element(ElementType.TEXT, "hello")])
        JSONExporter().export_document(doc, path)
        data = json.loads(path.read_text(encoding="utf-8"))
        assert len(data["documents"]) == 1


# ── ExcelExporter ──────────────────────────────────────────────────────────────

class TestExcelExporter:
    @pytest.fixture(autouse=True)
    def skip_without_openpyxl(self):
        pytest.importorskip("openpyxl", reason="openpyxl required for ExcelExporter")

    def test_sheets_created(self, tmp_path):
        import openpyxl
        from exporters import ExcelExporter
        path = tmp_path / "out.xlsx"
        elems = [make_element(ElementType.TEXT, "hello")]
        batch = make_batch([elems])
        ExcelExporter().export(batch, path)
        wb = openpyxl.load_workbook(path)
        assert "Summary" in wb.sheetnames
        assert "Documents" in wb.sheetnames
        assert "Elements" in wb.sheetnames

    def test_detail_sheets_created_when_enabled(self, tmp_path):
        import openpyxl
        from exporters import ExcelExporter
        path = tmp_path / "out.xlsx"
        batch = make_batch([[_table_elem(), _list_elem()]])
        cfg = ExporterConfig(excel_tables_sheet=True, excel_lists_sheet=True)
        ExcelExporter(cfg).export(batch, path)
        wb = openpyxl.load_workbook(path)
        assert "Tables" in wb.sheetnames
        assert "Lists" in wb.sheetnames

    def test_detail_sheets_suppressed_when_disabled(self, tmp_path):
        import openpyxl
        from exporters import ExcelExporter
        path = tmp_path / "out.xlsx"
        batch = make_batch([[make_element(ElementType.TEXT, "x")]])
        cfg = ExporterConfig(excel_tables_sheet=False, excel_lists_sheet=False)
        ExcelExporter(cfg).export(batch, path)
        wb = openpyxl.load_workbook(path)
        assert "Tables" not in wb.sheetnames
        assert "Lists" not in wb.sheetnames

    def test_elements_row_count(self, tmp_path):
        import openpyxl
        from exporters import ExcelExporter
        path = tmp_path / "out.xlsx"
        elems = [make_element(ElementType.TEXT, f"p{i}") for i in range(5)]
        batch = make_batch([elems])
        ExcelExporter().export(batch, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Elements"]
        data_rows = ws.max_row - 1  # subtract header row
        assert data_rows == 5

    def test_tables_sheet_explodes_table_elements(self, tmp_path):
        import openpyxl
        from exporters import ExcelExporter
        path = tmp_path / "out.xlsx"
        batch = make_batch([[_table_elem()]])
        ExcelExporter().export(batch, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Tables"]
        assert ws.max_row == 2  # header + 1 table row

    def test_lists_sheet_explodes_list_items(self, tmp_path):
        import openpyxl
        from exporters import ExcelExporter
        path = tmp_path / "out.xlsx"
        batch = make_batch([[_list_elem()]])
        ExcelExporter().export(batch, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Lists"]
        # _list_elem has 3 items → 3 rows + header
        assert ws.max_row == 4

    def test_no_openpyxl_raises_import_error(self):
        import unittest.mock as mock
        import sys
        # Temporarily hide openpyxl from sys.modules
        real_openpyxl = sys.modules.get("openpyxl")
        sys.modules["openpyxl"] = None  # type: ignore
        try:
            # Re-importing exporters won't help since the module is already loaded;
            # instead test by patching _OPENPYXL_AVAILABLE
            import exporters as exp
            original = exp._OPENPYXL_AVAILABLE
            exp._OPENPYXL_AVAILABLE = False
            with pytest.raises(ImportError):
                exp.ExcelExporter()
            exp._OPENPYXL_AVAILABLE = original
        finally:
            if real_openpyxl is not None:
                sys.modules["openpyxl"] = real_openpyxl
            elif "openpyxl" in sys.modules:
                del sys.modules["openpyxl"]
