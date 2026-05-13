"""
Export Pipeline

Converts BatchResult / DocumentResult objects to flat CSV, hierarchical JSON,
and multi-sheet Excel files.

Three exporter classes follow a common interface:
    exporter.export(batch_result, path)   → writes the file at *path*
    exporter.export_document(doc, path)   → convenience wrapper for a single doc

Design principles
-----------------
- Faithful: content is taken from OCR output stored in the model; nothing is
  synthesized.  For non-text elements, a bracketed descriptor (e.g. "[Table 3×4]")
  is used for the human-readable text column.
- Robust: every type-specific field extraction is wrapped in a try/except so a
  single malformed element never aborts the whole export.
- Parameterisable: all behaviour is controlled via ExporterConfig fields.
- Optional dependencies: openpyxl (Excel) is imported lazily; CSVExporter and
  JSONExporter require only the standard library.

Usage
-----
    from exporters import CSVExporter, JSONExporter, ExcelExporter, ExporterConfig

    cfg = ExporterConfig(include_raw_text=True)
    CSVExporter(cfg).export(batch, "out/results.csv")
    JSONExporter(cfg).export(batch, "out/results.json")
    ExcelExporter(cfg).export(batch, "out/results.xlsx")   # needs openpyxl
"""

import csv
import json
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from data_models import (
    Annotation,
    Barcode,
    BatchResult,
    BlockQuote,
    Caption,
    CodeBlock,
    DocumentResult,
    ElementType,
    EquationReference,
    FigureRegion,
    FormulaExpression,
    IndexEntry,
    ListStructure,
    PageFooter,
    PageHeader,
    ProcessingStatus,
    Reference,
    StructuralElement,
    TableOfContents,
    TableStructure,
    Watermark,
)

logger = logging.getLogger(__name__)


# ── Config ─────────────────────────────────────────────────────────────────────

@dataclass
class ExporterConfig:
    """Shared configuration for all exporter classes."""

    # CSV / Excel: include type-specific detail columns
    include_type_columns: bool = True

    # JSON: emit hierarchy (children embedded) vs flat list
    json_hierarchical: bool = True

    # JSON: indent for pretty-printing (None → compact)
    json_indent: Optional[int] = 2

    # Excel: include a "Tables" detail sheet
    excel_tables_sheet: bool = True

    # Excel: include a "Lists" detail sheet
    excel_lists_sheet: bool = True

    # Encoding for text files
    encoding: str = "utf-8"


# ── Content serialisation helpers ─────────────────────────────────────────────

def _content_text(elem: StructuralElement) -> str:
    """Return a flat text representation of an element's content."""
    c = elem.content
    try:
        if isinstance(c, str):
            return c
        if isinstance(c, TableStructure):
            return f"[Table {c.num_rows}×{c.num_cols}]"
        if isinstance(c, ListStructure):
            return " | ".join(item.content for item in c.items)
        if isinstance(c, FormulaExpression):
            return c.latex or c.raw_text
        if isinstance(c, EquationReference):
            return c.formula.latex or c.formula.raw_text
        if isinstance(c, Annotation):
            return c.content
        if isinstance(c, Barcode):
            return c.decoded_value
        if isinstance(c, Reference):
            return c.content
        if isinstance(c, CodeBlock):
            return c.content
        if isinstance(c, Watermark):
            return c.content or ""
        if isinstance(c, FigureRegion):
            return c.extracted_text or f"[Figure:{c.figure_type}]"
        if isinstance(c, TableOfContents):
            return c.title
        if isinstance(c, IndexEntry):
            return c.term
        if isinstance(c, PageHeader):
            return c.content
        if isinstance(c, PageFooter):
            return c.content
        if isinstance(c, Caption):
            return c.content
        if isinstance(c, BlockQuote):
            return c.content
        return str(c)
    except Exception:
        return ""


def _type_columns(elem: StructuralElement) -> Dict[str, Any]:
    """Return type-specific column dict (empty strings for absent fields)."""
    empty: Dict[str, Any] = {
        "table_rows": "", "table_cols": "",
        "table_has_headers": "", "table_has_merged_cells": "",
        "list_type": "", "list_item_count": "", "list_max_level": "",
        "formula_latex": "", "equation_number": "",
        "annotation_type": "", "annotation_color": "",
        "ref_type": "", "ref_id": "", "ref_location": "", "ref_target": "",
        "barcode_type": "", "barcode_value": "",
        "code_language": "",
        "toc_page_ref": "", "toc_level": "", "toc_target_heading": "",
        "watermark_opacity": "",
        "figure_type": "", "figure_caption_id": "",
        "index_term": "", "index_pages": "", "index_see_also": "",
    }
    c = elem.content
    try:
        if isinstance(c, TableStructure):
            empty["table_rows"] = c.num_rows
            empty["table_cols"] = c.num_cols
            empty["table_has_headers"] = bool(c.headers)
            empty["table_has_merged_cells"] = c.has_irregular_structure
        elif isinstance(c, ListStructure):
            empty["list_type"] = c.list_type
            empty["list_item_count"] = len(c.items)
            empty["list_max_level"] = max((i.level for i in c.items), default=0)
        elif isinstance(c, FormulaExpression):
            empty["formula_latex"] = c.latex or ""
        elif isinstance(c, EquationReference):
            empty["formula_latex"] = c.formula.latex or ""
            empty["equation_number"] = c.equation_number
        elif isinstance(c, Annotation):
            empty["annotation_type"] = c.annotation_type
            empty["annotation_color"] = c.color or ""
        elif isinstance(c, Barcode):
            empty["barcode_type"] = c.barcode_type
            empty["barcode_value"] = c.decoded_value
        elif isinstance(c, Reference):
            empty["ref_type"] = c.ref_type
            empty["ref_id"] = c.reference_id
            empty["ref_location"] = c.location
            empty["ref_target"] = c.target_ref or ""
        elif isinstance(c, CodeBlock):
            empty["code_language"] = c.language or ""
        elif isinstance(c, TableOfContents):
            empty["toc_page_ref"] = c.page_number
            empty["toc_level"] = c.level
            empty["toc_target_heading"] = c.target_heading_id or ""
        elif isinstance(c, Watermark):
            empty["watermark_opacity"] = c.opacity_estimate if c.opacity_estimate is not None else ""
        elif isinstance(c, FigureRegion):
            empty["figure_type"] = c.figure_type
            empty["figure_caption_id"] = c.caption_id or ""
        elif isinstance(c, IndexEntry):
            empty["index_term"] = c.term
            empty["index_pages"] = ",".join(str(p) for p in c.page_numbers)
            empty["index_see_also"] = "; ".join(c.see_also)
    except Exception:
        pass
    return empty


def _elem_base_row(
    batch_id: str,
    doc: DocumentResult,
    elem: StructuralElement,
) -> Dict[str, Any]:
    """Build the common columns for one element row."""
    return {
        "batch_id": batch_id,
        "document_id": doc.metadata.document_id,
        "source_file": doc.metadata.source_file,
        "element_id": elem.element_id,
        "element_type": elem.element_type.value,
        "page": elem.page_number,
        "nesting_level": elem.nesting_level,
        "parent_id": elem.parent_id or "",
        "child_count": len(elem.child_ids),
        "bbox_x_min": elem.bbox.x_min,
        "bbox_y_min": elem.bbox.y_min,
        "bbox_x_max": elem.bbox.x_max,
        "bbox_y_max": elem.bbox.y_max,
        "confidence": round(elem.confidence, 4),
        "processing_method": elem.processing_method or "",
        "content_text": _content_text(elem),
    }


def _all_elem_rows(
    batch: BatchResult,
    include_type_columns: bool,
) -> List[Dict[str, Any]]:
    """Produce one flat dict per element across the whole batch."""
    rows = []
    for doc in batch.documents:
        for elem in doc.elements:
            row = _elem_base_row(batch.batch_id, doc, elem)
            if include_type_columns:
                row.update(_type_columns(elem))
            rows.append(row)
    return rows


# ── CSV Exporter ───────────────────────────────────────────────────────────────

class CSVExporter:
    """
    Writes a flat CSV with one row per StructuralElement.

    Columns: common fields + (optionally) type-specific detail columns.
    The file is UTF-8 encoded with a BOM so Excel opens it correctly on Windows.
    """

    def __init__(self, config: Optional[ExporterConfig] = None) -> None:
        self.config = config or ExporterConfig()

    def export(self, batch: BatchResult, path: Union[str, Path]) -> None:
        """Write batch to CSV at *path*."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        rows = _all_elem_rows(batch, self.config.include_type_columns)
        if not rows:
            logger.warning("CSVExporter: no elements to export")
            path.write_text("", encoding=self.config.encoding)
            return
        fieldnames = list(rows[0].keys())
        with open(path, "w", newline="", encoding=self.config.encoding,
                  errors="replace") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        logger.info("CSVExporter: wrote %d rows to %s", len(rows), path)

    def export_document(self, doc: DocumentResult, path: Union[str, Path]) -> None:
        """Convenience wrapper: export a single DocumentResult."""
        from data_models import BatchResult as BR
        dummy = BR(
            batch_id=doc.metadata.document_id,
            created_at=doc.metadata.processing_timestamp,
            documents=[doc],
        )
        self.export(dummy, path)


# ── JSON Exporter ──────────────────────────────────────────────────────────────

def _elem_to_dict(elem: StructuralElement, include_type_columns: bool) -> Dict[str, Any]:
    """Serialise a StructuralElement to a plain dict for JSON output."""
    d: Dict[str, Any] = {
        "element_id": elem.element_id,
        "element_type": elem.element_type.value,
        "page": elem.page_number,
        "nesting_level": elem.nesting_level,
        "parent_id": elem.parent_id,
        "child_ids": elem.child_ids,
        "bbox": {
            "x_min": elem.bbox.x_min,
            "y_min": elem.bbox.y_min,
            "x_max": elem.bbox.x_max,
            "y_max": elem.bbox.y_max,
        },
        "confidence": round(elem.confidence, 4),
        "processing_method": elem.processing_method,
        "content_text": _content_text(elem),
        "metadata": elem.metadata,
    }
    if include_type_columns:
        d["type_detail"] = {k: v for k, v in _type_columns(elem).items() if v != ""}
    return d


def _doc_to_dict(
    doc: DocumentResult,
    hierarchical: bool,
    include_type_columns: bool,
) -> Dict[str, Any]:
    """Serialise a DocumentResult."""
    meta = doc.metadata
    doc_dict: Dict[str, Any] = {
        "document_id": meta.document_id,
        "source_file": meta.source_file,
        "processing_timestamp": meta.processing_timestamp.isoformat(),
        "processing_duration_s": round(meta.processing_duration, 4),
        "image_dimensions": list(meta.image_dimensions),
        "detected_language": meta.detected_language,
        "total_elements": meta.total_elements_extracted,
        "average_confidence": round(meta.average_confidence, 4),
        "processing_status": meta.processing_status.value,
        "errors": meta.errors_encountered,
    }

    elem_dicts = {
        e.element_id: _elem_to_dict(e, include_type_columns)
        for e in doc.elements
    }

    if hierarchical:
        # Embed children inside their parents; collect root nodes
        for elem in doc.elements:
            ed = elem_dicts[elem.element_id]
            ed["children"] = []  # will be populated below

        for elem in doc.elements:
            if elem.parent_id and elem.parent_id in elem_dicts:
                elem_dicts[elem.parent_id]["children"].append(
                    elem_dicts[elem.element_id]
                )

        roots = [
            elem_dicts[e.element_id]
            for e in doc.elements
            if e.parent_id is None or e.parent_id not in elem_dicts
        ]
        doc_dict["elements"] = roots
    else:
        doc_dict["elements"] = list(elem_dicts.values())

    return doc_dict


class JSONExporter:
    """
    Writes a JSON file containing one object per document.

    In hierarchical mode (default), child elements are embedded inside
    their parents to reflect the document structure.  In flat mode,
    all elements are listed in a top-level array.
    """

    def __init__(self, config: Optional[ExporterConfig] = None) -> None:
        self.config = config or ExporterConfig()

    def export(self, batch: BatchResult, path: Union[str, Path]) -> None:
        """Write batch to JSON at *path*."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        cfg = self.config
        stats = batch.statistics
        output: Dict[str, Any] = {
            "batch_id": batch.batch_id,
            "created_at": batch.created_at.isoformat(),
            "statistics": {
                "total_documents": stats.total_documents if stats else len(batch.documents),
                "successful": stats.successful_documents if stats else 0,
                "failed": stats.failed_documents if stats else 0,
                "total_elements": stats.total_elements if stats else 0,
                "average_confidence": round(stats.average_confidence, 4) if stats else 0.0,
                "total_processing_time_s": round(stats.total_processing_time, 4) if stats else 0.0,
            },
            "documents": [
                _doc_to_dict(doc, cfg.json_hierarchical, cfg.include_type_columns)
                for doc in batch.documents
            ],
        }
        with open(path, "w", encoding=cfg.encoding, errors="replace") as fh:
            json.dump(output, fh, indent=cfg.json_indent, ensure_ascii=False,
                      default=str)
        total_elems = sum(len(d.elements) for d in batch.documents)
        logger.info("JSONExporter: wrote %d docs / %d elements to %s",
                    len(batch.documents), total_elems, path)

    def export_document(self, doc: DocumentResult, path: Union[str, Path]) -> None:
        """Convenience wrapper: export a single DocumentResult."""
        from data_models import BatchResult as BR
        dummy = BR(
            batch_id=doc.metadata.document_id,
            created_at=doc.metadata.processing_timestamp,
            documents=[doc],
        )
        self.export(dummy, path)


# ── Excel Exporter ─────────────────────────────────────────────────────────────

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    import pandas as _pd  # noqa: F401
    _PANDAS_AVAILABLE = True
except ImportError:
    _PANDAS_AVAILABLE = False


# ── DataFrame Exporter ─────────────────────────────────────────────────────────

class DataFrameExporter:
    """
    Converts BatchResult / DocumentResult to a pandas DataFrame.

    Each row corresponds to one StructuralElement.  Columns match those
    produced by CSVExporter (common fields + optional type-specific detail).

    Optionally serialises the DataFrame to a pickle (.pkl) file so the result
    can be reloaded in a subsequent session without re-running OCR.

    Requires pandas.  Raises ImportError at construction time if not installed.

    Usage::

        exporter = DataFrameExporter()

        # In-memory DataFrame
        df = exporter.to_dataframe(batch)

        # Save to pickle and get the DataFrame back
        df = exporter.export(batch, "results/batch.pkl")

        # Single-document convenience
        df = exporter.export_document(doc, "results/page1.pkl")
    """

    def __init__(self, config: Optional[ExporterConfig] = None) -> None:
        if not _PANDAS_AVAILABLE:
            raise ImportError(
                "pandas is required for DataFrameExporter. "
                "Install it with: pip install pandas"
            )
        self.config = config or ExporterConfig()

    def to_dataframe(self, batch: BatchResult) -> "Any":  # -> pd.DataFrame
        """Return a pandas DataFrame with one row per element in *batch*."""
        import pandas as pd
        rows = _all_elem_rows(batch, self.config.include_type_columns)
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows)

    def export(
        self,
        batch: BatchResult,
        path: Union[str, Path],
    ) -> "Any":  # -> pd.DataFrame
        """
        Convert *batch* to a DataFrame and save it as a pickle file at *path*.

        Returns the DataFrame so callers can use it immediately without
        loading the file.
        """
        import pandas as pd
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        df = self.to_dataframe(batch)
        df.to_pickle(path)
        logger.info("DataFrameExporter: wrote %d rows to %s", len(df), path)
        return df

    def export_document(
        self,
        doc: DocumentResult,
        path: Optional[Union[str, Path]] = None,
    ) -> "Any":  # -> pd.DataFrame
        """
        Convenience wrapper: convert a single DocumentResult to a DataFrame.

        If *path* is given, also saves to that pickle file.
        """
        dummy = BatchResult(
            batch_id=doc.metadata.document_id,
            created_at=doc.metadata.processing_timestamp,
            documents=[doc],
        )
        if path is not None:
            return self.export(dummy, path)
        return self.to_dataframe(dummy)


class ExcelExporter:
    """
    Writes a multi-sheet .xlsx file.

    Sheets
    ------
    Summary   — batch-level metadata and statistics
    Documents — one row per DocumentResult
    Elements  — flat one-row-per-element (same columns as CSVExporter)
    Tables    — one row per TABLE element with 2D cell data (optional)
    Lists     — one row per LIST element with item text (optional)

    Requires openpyxl.  Raises ImportError if not installed.
    """

    _HEADER_FILL = "4472C4"   # Excel blue
    _HEADER_FONT_COLOR = "FFFFFF"
    _ALT_ROW_FILL = "DCE6F1"

    def __init__(self, config: Optional[ExporterConfig] = None) -> None:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for ExcelExporter. "
                "Install it with: pip install openpyxl"
            )
        self.config = config or ExporterConfig()

    # ── Public API ─────────────────────────────────────────────────────────────

    def export(self, batch: BatchResult, path: Union[str, Path]) -> None:
        """Write batch to .xlsx at *path*."""
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        self._write_summary(wb, batch)
        self._write_documents(wb, batch)
        self._write_elements(wb, batch)
        if self.config.excel_tables_sheet:
            self._write_tables(wb, batch)
        if self.config.excel_lists_sheet:
            self._write_lists(wb, batch)

        wb.save(path)
        total_elems = sum(len(d.elements) for d in batch.documents)
        logger.info("ExcelExporter: wrote %d docs / %d elements to %s",
                    len(batch.documents), total_elems, path)

    def export_document(self, doc: DocumentResult, path: Union[str, Path]) -> None:
        """Convenience wrapper: export a single DocumentResult."""
        from data_models import BatchResult as BR
        dummy = BR(
            batch_id=doc.metadata.document_id,
            created_at=doc.metadata.processing_timestamp,
            documents=[doc],
        )
        self.export(dummy, path)

    # ── Sheet writers ──────────────────────────────────────────────────────────

    def _write_summary(self, wb: Any, batch: BatchResult) -> None:
        ws = wb.create_sheet("Summary")
        stats = batch.statistics
        rows = [
            ("Batch ID", batch.batch_id),
            ("Created At", batch.created_at.isoformat()),
            ("Total Documents", stats.total_documents if stats else len(batch.documents)),
            ("Successful Documents", stats.successful_documents if stats else ""),
            ("Failed Documents", stats.failed_documents if stats else ""),
            ("Partial Documents", stats.partial_documents if stats else ""),
            ("Total Elements Extracted", stats.total_elements if stats else ""),
            ("Average Confidence", round(stats.average_confidence, 4) if stats else ""),
            ("Total Processing Time (s)", round(stats.total_processing_time, 4) if stats else ""),
            ("Avg Time per Document (s)", round(stats.average_time_per_document, 4) if stats else ""),
        ]
        if stats and stats.elements_by_type:
            rows.append(("", ""))
            rows.append(("Elements by Type", ""))
            for etype, count in sorted(stats.elements_by_type.items(),
                                       key=lambda x: x[1], reverse=True):
                rows.append((f"  {etype.value}", count))

        for key, val in rows:
            ws.append([key, val])

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 24
        self._style_header_row(ws, 1)

    def _write_documents(self, wb: Any, batch: BatchResult) -> None:
        ws = wb.create_sheet("Documents")
        headers = [
            "batch_id", "document_id", "source_file",
            "processing_timestamp", "processing_duration_s",
            "image_width", "image_height",
            "detected_language", "total_elements",
            "average_confidence", "processing_status",
            "errors",
        ]
        ws.append(headers)
        self._style_header_row(ws, 1)

        for i, doc in enumerate(batch.documents, start=2):
            m = doc.metadata
            w, h = m.image_dimensions
            ws.append([
                batch.batch_id,
                m.document_id,
                m.source_file,
                m.processing_timestamp.isoformat(),
                round(m.processing_duration, 4),
                w, h,
                m.detected_language,
                m.total_elements_extracted,
                round(m.average_confidence, 4),
                m.processing_status.value,
                "; ".join(m.errors_encountered),
            ])
            if i % 2 == 0:
                self._fill_row(ws, i, self._ALT_ROW_FILL)

        self._autofit_columns(ws)

    def _write_elements(self, wb: Any, batch: BatchResult) -> None:
        ws = wb.create_sheet("Elements")
        rows = _all_elem_rows(batch, self.config.include_type_columns)
        if not rows:
            ws.append(["No elements"])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        self._style_header_row(ws, 1)
        for i, row in enumerate(rows, start=2):
            ws.append([row.get(h, "") for h in headers])
            if i % 2 == 0:
                self._fill_row(ws, i, self._ALT_ROW_FILL)
        self._autofit_columns(ws)

    def _write_tables(self, wb: Any, batch: BatchResult) -> None:
        """One row per TABLE element; embedded 2D cell data as serialised text."""
        ws = wb.create_sheet("Tables")
        headers = [
            "batch_id", "document_id", "element_id", "page",
            "num_rows", "num_cols", "has_headers", "has_merged_cells",
            "confidence", "bbox",
            "cells_csv",
        ]
        ws.append(headers)
        self._style_header_row(ws, 1)

        row_idx = 2
        for doc in batch.documents:
            for elem in doc.elements:
                if elem.element_type != ElementType.TABLE:
                    continue
                c = elem.content
                if not isinstance(c, TableStructure):
                    continue
                cells_csv = c.to_csv()
                bbox_str = (f"{elem.bbox.x_min},{elem.bbox.y_min},"
                            f"{elem.bbox.x_max},{elem.bbox.y_max}")
                ws.append([
                    batch.batch_id,
                    doc.metadata.document_id,
                    elem.element_id,
                    elem.page_number,
                    c.num_rows,
                    c.num_cols,
                    bool(c.headers),
                    c.has_irregular_structure,
                    round(elem.confidence, 4),
                    bbox_str,
                    cells_csv,
                ])
                if row_idx % 2 == 0:
                    self._fill_row(ws, row_idx, self._ALT_ROW_FILL)
                row_idx += 1

        self._autofit_columns(ws)

    def _write_lists(self, wb: Any, batch: BatchResult) -> None:
        """One row per LIST item (exploded), preserving level and type."""
        ws = wb.create_sheet("Lists")
        headers = [
            "batch_id", "document_id", "list_element_id", "page",
            "list_type", "item_index", "item_level",
            "item_list_type", "item_content", "confidence",
        ]
        ws.append(headers)
        self._style_header_row(ws, 1)

        row_idx = 2
        for doc in batch.documents:
            for elem in doc.elements:
                if elem.element_type != ElementType.LIST:
                    continue
                c = elem.content
                if not isinstance(c, ListStructure):
                    continue
                for idx, item in enumerate(c.items):
                    ws.append([
                        batch.batch_id,
                        doc.metadata.document_id,
                        elem.element_id,
                        elem.page_number,
                        c.list_type,
                        idx,
                        item.level,
                        item.list_type,
                        item.content,
                        round(item.confidence, 4),
                    ])
                    if row_idx % 2 == 0:
                        self._fill_row(ws, row_idx, self._ALT_ROW_FILL)
                    row_idx += 1

        self._autofit_columns(ws)

    # ── Formatting helpers ─────────────────────────────────────────────────────

    def _style_header_row(self, ws: Any, row: int) -> None:
        fill = PatternFill("solid", fgColor=self._HEADER_FILL)
        font = Font(bold=True, color=self._HEADER_FONT_COLOR)
        for cell in ws[row]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=False)

    def _fill_row(self, ws: Any, row: int, hex_color: str) -> None:
        fill = PatternFill("solid", fgColor=hex_color)
        for cell in ws[row]:
            cell.fill = fill

    @staticmethod
    def _autofit_columns(ws: Any, max_width: int = 60) -> None:
        for col in ws.columns:
            best = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(best + 2, max_width)
